from io import BytesIO
from openpyxl import load_workbook
from app.schemas.points import PointsImportPreview

def parse_excel_file(file_bytes: bytes) -> list[PointsImportPreview]:
    try:
        wb = load_workbook(filename=BytesIO(file_bytes))
        ws = wb.active
        results = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue
            
            student_id = row[0]
            name = str(row[1]) if row[1] else ""
            class_name = str(row[2]) if row[2] else ""
            change_amount_raw = row[3]
            reason = str(row[4]) if row[4] else ""
            
            valid = True
            error = None
            
            try:
                if student_id is not None:
                    student_id = int(student_id)
                change_amount = int(change_amount_raw)
            except (ValueError, TypeError):
                valid = False
                error = "数据格式错误"
            
            if not name or not class_name:
                valid = False
                error = "姓名或班级为空"
            
            results.append(PointsImportPreview(
                row=row_idx,
                student_id=student_id if isinstance(student_id, int) else None,
                name=name,
                class_name=class_name,
                change_amount=change_amount if 'change_amount' in locals() else 0,
                reason=reason,
                valid=valid,
                error=error
            ))
        
        return results
    except Exception as e:
        raise ValueError(f"解析Excel文件失败: {str(e)}")
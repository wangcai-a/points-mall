from sqlalchemy.orm import Session
from io import BytesIO
from openpyxl import load_workbook
from app.models import Student, PointsRecord, Teacher
from app.schemas.student import StudentCreate, StudentUpdate, StudentResponse, PointsHistoryResponse, StudentImportPreview, StudentImportResponse
from app.database import Base, engine

def get_students(db: Session, page: int = 1, page_size: int = 10, name: str = None, class_name: str = None):
    query = db.query(Student)
    
    if name:
        query = query.filter(Student.name.ilike(f"%{name}%"))
    if class_name:
        query = query.filter(Student.class_name == class_name)
    
    total = query.count()
    students = query.offset((page - 1) * page_size).limit(page_size).all()
    
    return {
        "list": [StudentResponse.model_validate(s) for s in students],
        "total": total,
        "page": page,
        "page_size": page_size
    }

def get_student(db: Session, student_id: int) -> Student | None:
    return db.query(Student).filter(Student.id == student_id).first()

def get_class_list(db: Session) -> list[str]:
    result = db.query(Student.class_name).distinct().all()
    return [row[0] for row in result if row[0]]

def create_student(db: Session, student: StudentCreate) -> Student:
    db_student = Student(
        name=student.name,
        class_name=student.class_name,
        total_points=student.total_points
    )
    db.add(db_student)
    db.commit()
    db.refresh(db_student)
    return db_student

def update_student(db: Session, student_id: int, student_update: StudentUpdate) -> Student | None:
    db_student = db.query(Student).filter(Student.id == student_id).first()
    if db_student:
        db_student.name = student_update.name
        db_student.class_name = student_update.class_name
        db_student.updated_at = __import__('sqlalchemy').sql.func.current_timestamp()
        db.commit()
        db.refresh(db_student)
    return db_student

def delete_student(db: Session, student_id: int) -> bool:
    db_student = db.query(Student).filter(Student.id == student_id).first()
    if db_student:
        db.delete(db_student)
        db.commit()
        return True
    return False

def get_points_history(db: Session, student_id: int, page: int = 1, page_size: int = 10):
    query = db.query(PointsRecord, Teacher.name.label('teacher_name')).\
        join(Teacher, PointsRecord.teacher_id == Teacher.id).\
        filter(PointsRecord.student_id == student_id).\
        order_by(PointsRecord.created_at.desc())
    
    total = query.count()
    records = query.offset((page - 1) * page_size).limit(page_size).all()
    
    return {
        "list": [
            PointsHistoryResponse(
                id=record[0].id,
                change_amount=record[0].change_amount,
                reason=record[0].reason,
                type=record[0].type,
                teacher_name=record[1],
                created_at=record[0].created_at
            )
            for record in records
        ],
        "total": total
    }

def parse_student_excel(file_bytes: bytes) -> list[StudentImportPreview]:
    try:
        wb = load_workbook(filename=BytesIO(file_bytes))
        ws = wb.active
        results = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue
            
            name = str(row[0]).strip() if row[0] else ""
            class_name = str(row[1]).strip() if row[1] else ""
            total_points = row[2] if row[2] else 0
            
            valid = True
            error = None
            
            if not name:
                valid = False
                error = "姓名为空"
            
            if not class_name:
                valid = False
                error = "班级为空"
            
            try:
                total_points = int(total_points)
            except (ValueError, TypeError):
                valid = False
                error = "积分格式错误"
            
            results.append(StudentImportPreview(
                row=row_idx,
                name=name,
                class_name=class_name,
                total_points=total_points if isinstance(total_points, int) else 0,
                valid=valid,
                error=error
            ))
        
        return results
    except Exception as e:
        raise ValueError(f"解析Excel文件失败: {str(e)}")

def import_students(db: Session, students_data: list[dict]) -> StudentImportResponse:
    success_count = 0
    fail_count = 0
    
    for data in students_data:
        try:
            existing_student = db.query(Student).filter(
                Student.name == data['name'],
                Student.class_name == data['class_name']
            ).first()
            
            if existing_student:
                fail_count += 1
                continue
            
            db_student = Student(
                name=data['name'],
                class_name=data['class_name'],
                total_points=data.get('total_points', 0)
            )
            db.add(db_student)
            success_count += 1
        except Exception:
            fail_count += 1
    
    db.commit()
    
    return StudentImportResponse(
        success_count=success_count,
        fail_count=fail_count
    )